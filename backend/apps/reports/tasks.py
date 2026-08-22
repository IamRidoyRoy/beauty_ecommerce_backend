import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from celery import shared_task
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import ReportExport
from .selectors import REPORTS


def _excel_value(value):
    if value is None:
        return ""
    # Excel/OpenPyXL rejects timezone-aware datetimes; strings are portable.
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=str, ensure_ascii=False)
    if isinstance(value, Decimal):
        return float(value)
    return value


@shared_task
def export_report(export_id):
    export = ReportExport.objects.get(pk=export_id)
    export.status = ReportExport.Status.PROCESSING
    export.save(update_fields=["status", "updated_at"])
    try:
        result = REPORTS[export.report](export.params)
        rows = result if isinstance(result, list) else [result]
        keys = []
        for row in rows:
            if isinstance(row, dict):
                for key in row:
                    if key not in keys:
                        keys.append(key)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        if keys:
            ws.append(keys)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append([_excel_value(row.get(key, "")) for key in keys])
            for column in ws.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = width
        else:
            ws["A1"] = "No data for the selected report range."

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        export.file.save(f"{export.report}-{export.id}.xlsx", ContentFile(out.read()), save=False)
        export.status = ReportExport.Status.COMPLETED
        export.error = ""
        export.save(update_fields=["file", "status", "error", "updated_at"])
    except Exception as exc:
        export.status = ReportExport.Status.FAILED
        export.error = str(exc)
        export.save(update_fields=["status", "error", "updated_at"])
        raise
    return export.id
