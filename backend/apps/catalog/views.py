import csv
from io import TextIOWrapper, BytesIO
from decimal import Decimal, InvalidOperation
from django.utils.text import slugify
from django.utils import timezone
from django.http import HttpResponse
from django.db.models.functions import Coalesce
from django.db import transaction
from django.db.models import Q, Min, Max, Sum, Value, IntegerField, Case, When
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser,FormParser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.permissions import AllowAny,IsAuthenticated
from rest_framework.viewsets import ReadOnlyModelViewSet,ModelViewSet
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError
from apps.accounts.models import UserRole
from apps.common.permissions import role_permission
from apps.common.responses import success
from .models import *
from .serializers import *
from .selectors import product_list_queryset,product_detail_queryset
from .filters import ProductFilter
from .services import publish_product,set_primary_image,reorder_images

CatalogAdmin=role_permission(UserRole.SUPER_ADMIN,UserRole.ADMIN,UserRole.MANAGER,UserRole.PRODUCT_MANAGER)
CatalogOrderRead=role_permission(UserRole.SUPER_ADMIN,UserRole.ADMIN,UserRole.MANAGER,UserRole.PRODUCT_MANAGER,UserRole.ORDER_MANAGER)
class ProductViewSet(ReadOnlyModelViewSet):
    permission_classes=[AllowAny]; lookup_field="slug"; filterset_class=ProductFilter; search_fields=("name","sku","brand__name","category__name"); ordering_fields=("created_at","base_price","name")
    def get_queryset(self): return product_detail_queryset() if self.action=="retrieve" else product_list_queryset()
    def get_serializer_class(self): return ProductDetailSerializer if self.action=="retrieve" else ProductListSerializer
    @action(detail=False,methods=["get"],url_path="search")
    def search(self,request): return self.list(request)

    @action(detail=False,methods=["get"],url_path="filter-meta")
    def filter_meta(self,request):
        """Return lightweight catalog bounds used by storefront filter controls."""
        qs=Product.objects.filter(status=Product.Status.ACTIVE)
        brand=request.query_params.get("brand")
        category=request.query_params.get("category")
        product_type=request.query_params.get("product_type")
        if brand: qs=qs.filter(brand__slug=brand)
        if category: qs=qs.filter(category__slug=category)
        if product_type: qs=qs.filter(product_type=product_type)
        bounds=qs.aggregate(min_price=Min("base_price"),max_price=Max("base_price"))
        min_price=bounds["min_price"] if bounds["min_price"] is not None else Decimal("0")
        max_price=bounds["max_price"] if bounds["max_price"] is not None else Decimal("10000")
        if max_price <= min_price: max_price=min_price+Decimal("1")
        return success({"min_price":float(min_price),"max_price":float(max_price)})
class CategoryViewSet(ReadOnlyModelViewSet):
    permission_classes=[AllowAny]; serializer_class=CategorySerializer
    queryset=(Category.objects.filter(active=True).select_related("parent")
        .annotate(_priority_bucket=Case(When(order=0, then=Value(1)), default=Value(0), output_field=IntegerField()))
        .order_by("_priority_bucket", "order", "name", "id"))
class BrandViewSet(ReadOnlyModelViewSet):
    permission_classes=[AllowAny]; serializer_class=BrandSerializer; queryset=Brand.objects.filter(active=True)

class WishlistView(APIView):
    permission_classes=[IsAuthenticated]
    def get(self,request): return success(ProductListSerializer(product_list_queryset().filter(wishlisted_by__user=request.user),many=True,context={"request":request}).data)
    def post(self,request):
        product=Product.objects.get(pk=request.data.get("product")); WishlistItem.objects.get_or_create(user=request.user,product=product)
        from apps.common.models import AnalyticsEvent
        AnalyticsEvent.objects.create(event_type=AnalyticsEvent.EventType.WISHLIST,user=request.user,product_id_ref=product.id)
        return success(message="Added to wishlist.",status=201)
    def delete(self,request): WishlistItem.objects.filter(user=request.user,product_id=request.data.get("product")).delete(); return success(message="Removed from wishlist.")

class AdminProductViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=ProductAdminSerializer
    def get_permissions(self):
        classes=[CatalogOrderRead] if self.request.method in {"GET","HEAD","OPTIONS"} else [CatalogAdmin]
        return [permission() for permission in classes]
    queryset=(Product.objects.all()
        .select_related("brand","category")
        .prefetch_related("variants","images")
        .annotate(
            simple_available_stock=Coalesce(Sum("stock_item__stocks__available_stock"),Value(0),output_field=IntegerField()),
            variant_available_stock=Coalesce(Sum("variants__stock_item__stocks__available_stock"),Value(0),output_field=IntegerField()),
        )
        .order_by("-id"))
    filterset_fields=("product_type","status","brand","category","featured","new_arrival","bestseller","trending"); search_fields=("name","sku","barcode","brand__name","category__name","variants__sku"); ordering_fields=("id","created_at","updated_at","base_price","name")
    @action(detail=True,methods=["post"])
    def publish(self,request,pk=None): return success(ProductAdminSerializer(publish_product(product=self.get_object())).data,"Product published.")
    @staticmethod
    def _product_excel_response(workbook, filename):
        output=BytesIO(); workbook.save(output); output.seek(0)
        response=HttpResponse(output.getvalue(),content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def _style_excel_sheet(ws, freeze="A2"):
        ws.freeze_panes=freeze
        ws.auto_filter.ref=ws.dimensions
        fill=PatternFill("solid",fgColor="111827")
        for cell in ws[1]:
            cell.font=Font(color="FFFFFF",bold=True); cell.fill=fill; cell.alignment=Alignment(vertical="center")
        widths={}
        for row in ws.iter_rows():
            for cell in row:
                value="" if cell.value is None else str(cell.value)
                widths[cell.column_letter]=min(max(widths.get(cell.column_letter,0),len(value)+2),42)
        for col,width in widths.items(): ws.column_dimensions[col].width=max(12,width)

    @action(detail=False,methods=["get"],url_path="import-template")
    def import_template(self,request):
        wb=Workbook(); products=wb.active; products.title="Products"
        product_headers=["product_id","name","slug","product_type","sku","barcode","brand","category","base_price","compare_at_price","cost_price","status","short_description","description","weight","tax_class","featured","new_arrival","bestseller","trending"]
        products.append(product_headers)
        products.append(["","Example Cleanser","","simple","EXAMPLE-SKU","","Example Brand","Skincare",1000,"",650,"draft","Short product description","Full product description",0.25,"","yes","yes","no","no"])
        products.append(["","Example Variable Product","","variable","","","Example Brand","Skincare",1500,"",900,"draft","","",0.3,"","no","no","no","no"])
        self._style_excel_sheet(products)
        type_validation=DataValidation(type="list",formula1='"simple,variable"',allow_blank=False); products.add_data_validation(type_validation); type_validation.add("D2:D5000")
        status_validation=DataValidation(type="list",formula1='"draft,active,archived"',allow_blank=False); products.add_data_validation(status_validation); status_validation.add("L2:L5000")
        bool_validation=DataValidation(type="list",formula1='"yes,no"',allow_blank=True); products.add_data_validation(bool_validation)
        for col in ("Q","R","S","T"): bool_validation.add(f"{col}2:{col}5000")

        variants=wb.create_sheet("Variants")
        variant_headers=["variant_id","product_id","product_slug","product_name","sku","barcode","price_override","cost_price","weight","is_active","attributes"]
        variants.append(variant_headers)
        variants.append(["","","example-variable-product","Example Variable Product","EXAMPLE-VAR-RED","",1550,920,0.3,"yes","Shade=Red | Size=50ml"])
        self._style_excel_sheet(variants)
        active_validation=DataValidation(type="list",formula1='"yes,no"',allow_blank=True); variants.add_data_validation(active_validation); active_validation.add("J2:J5000")

        refs=wb.create_sheet("References")
        refs.append(["Brands","Categories","Attribute","Attribute Value"])
        brands=list(Brand.objects.order_by("name").values_list("name",flat=True)); categories=list(Category.objects.order_by("name").values_list("name",flat=True))
        attr_values=list(AttributeValue.objects.select_related("attribute").order_by("attribute__name","value"))
        size=max(len(brands),len(categories),len(attr_values),1)
        for i in range(size):
            refs.append([brands[i] if i<len(brands) else "",categories[i] if i<len(categories) else "",attr_values[i].attribute.name if i<len(attr_values) else "",attr_values[i].value if i<len(attr_values) else ""])
        self._style_excel_sheet(refs)

        instructions=wb.create_sheet("Instructions",0)
        notes=[
            ["Bulk Product Import Instructions"],
            ["1. Use the Products sheet to create or update products. product_id is preferred for updates; simple products can also match by SKU."],
            ["2. Brand and Category must already exist; use their exact name, slug, or numeric ID."],
            ["3. For variable products, add variants in the Variants sheet. Match parent with product_id, product_slug, or exact product_name."],
            ["4. Variant attributes format: Shade=Red | Size=50ml. Attribute values must already exist."],
            ["5. Existing rows are updated; new rows are created. Blank optional cells clear nullable/text values only when explicitly provided by Excel."],
            ["6. Images are not imported from this workbook. Use Catalog > Product Images for image uploads."],
            ["7. Keep headers unchanged. Rows with errors are skipped and returned in the import result."],
        ]
        for row in notes: instructions.append(row)
        instructions.column_dimensions["A"].width=115; instructions["A1"].font=Font(bold=True,size=16); instructions["A1"].fill=PatternFill("solid",fgColor="DB2777"); instructions["A1"].font=Font(color="FFFFFF",bold=True,size=16)
        for row in instructions.iter_rows(): row[0].alignment=Alignment(wrap_text=True,vertical="top")
        return self._product_excel_response(wb,"product_import_template.xlsx")

    @action(detail=False,methods=["get"],url_path="export-file")
    def export_file(self,request):
        wb=Workbook(); ws=wb.active; ws.title="Products"
        headers=["product_id","uuid","name","slug","product_type","sku","barcode","brand","brand_id","category","category_id","base_price","compare_at_price","cost_price","status","short_description","description","weight","tax_class","featured","new_arrival","bestseller","trending","published_at","created_at","updated_at"]
        ws.append(headers)
        qs=Product.objects.select_related("brand","category").prefetch_related("variants__attributes__attribute","images").order_by("id")
        for p in qs:
            ws.append([p.id,str(p.uuid),p.name,p.slug,p.product_type,p.sku or "",p.barcode or "",p.brand.name,p.brand_id,p.category.name,p.category_id,float(p.base_price),float(p.compare_at_price) if p.compare_at_price is not None else "",float(p.cost_price) if p.cost_price is not None else "",p.status,p.short_description,p.description,float(p.weight) if p.weight is not None else "",p.tax_class,p.featured,p.new_arrival,p.bestseller,p.trending,p.published_at.isoformat() if p.published_at else "",p.created_at.isoformat() if p.created_at else "",p.updated_at.isoformat() if p.updated_at else ""])
        self._style_excel_sheet(ws)

        variants=wb.create_sheet("Variants"); variants.append(["variant_id","uuid","product_id","product_slug","product_name","sku","barcode","price_override","cost_price","weight","is_active","attributes","created_at","updated_at"])
        for v in ProductVariant.objects.select_related("product").prefetch_related("attributes__attribute").order_by("product_id","id"):
            attrs=" | ".join(f"{a.attribute.name}={a.value}" for a in v.attributes.all())
            variants.append([v.id,str(v.uuid),v.product_id,v.product.slug,v.product.name,v.sku,v.barcode or "",float(v.price_override) if v.price_override is not None else "",float(v.cost_price) if v.cost_price is not None else "",float(v.weight) if v.weight is not None else "",v.is_active,attrs,v.created_at.isoformat() if v.created_at else "",v.updated_at.isoformat() if v.updated_at else ""])
        self._style_excel_sheet(variants)

        images=wb.create_sheet("Images"); images.append(["image_id","product_id","product_name","variant_id","variant_sku","image_path","image_type","alt_text","order","is_primary"])
        for image in ProductImage.objects.select_related("product","variant").order_by("product_id","order","id"):
            images.append([image.id,image.product_id,image.product.name,image.variant_id or "",image.variant.sku if image.variant_id else "",image.image.url if image.image else "",image.image_type,image.alt_text,image.order,image.is_primary])
        self._style_excel_sheet(images)

        profiles=wb.create_sheet("Beauty Profiles"); profiles.append(["product_id","product_name","benefits","ingredients_text","how_to_use","precautions","country_of_origin","shelf_life","pao","skin_types","hair_types","concerns","ingredients"])
        for profile in ProductBeautyProfile.objects.select_related("product").prefetch_related("skin_types","hair_types","concerns","ingredients").order_by("product_id"):
            profiles.append([profile.product_id,profile.product.name,profile.benefits,profile.ingredients_text,profile.how_to_use,profile.precautions,profile.country_of_origin,profile.shelf_life,profile.pao," | ".join(x.name for x in profile.skin_types.all())," | ".join(x.name for x in profile.hair_types.all())," | ".join(x.name for x in profile.concerns.all())," | ".join(x.name for x in profile.ingredients.all())])
        self._style_excel_sheet(profiles)

        claims=wb.create_sheet("Product Claims"); claims.append(["product_claim_id","product_id","product_name","claim","is_verified","evidence","source_url","active","reviewed_at"])
        for pc in ProductClaim.objects.select_related("product","claim").order_by("product_id","id"):
            claims.append([pc.id,pc.product_id,pc.product.name,pc.claim.name,pc.is_verified,pc.evidence,pc.source_url,pc.active,pc.reviewed_at.isoformat() if pc.reviewed_at else ""])
        self._style_excel_sheet(claims)
        filename=f"products_export_{timezone.localdate().isoformat()}.xlsx"
        return self._product_excel_response(wb,filename)

    @action(detail=False,methods=["post"],url_path="import-file",parser_classes=[MultiPartParser,FormParser])
    def import_file(self,request):
        upload=request.FILES.get("file")
        if not upload: raise ValidationError({"file":"Upload a CSV or XLSX file."})
        lower=upload.name.lower()
        if not (lower.endswith(".csv") or lower.endswith(".xlsx")): raise ValidationError({"file":"Only .csv and .xlsx files are supported."})

        def sheet_rows(ws):
            values=list(ws.iter_rows(values_only=True))
            if not values:return []
            headers=[str(x or "").strip() for x in values[0]]
            return [dict(zip(headers,row)) for row in values[1:] if any(v not in (None,"") for v in row)]
        variant_rows=[]
        if lower.endswith(".csv"):
            wrapper=TextIOWrapper(upload.file,encoding="utf-8-sig",newline=""); product_rows=list(csv.DictReader(wrapper))
        else:
            wb=load_workbook(upload,read_only=True,data_only=True)
            products_ws=wb["Products"] if "Products" in wb.sheetnames else wb.active
            product_rows=sheet_rows(products_ws)
            if "Variants" in wb.sheetnames: variant_rows=sheet_rows(wb["Variants"])

        def has_key(row,key): return key in row and row.get(key) is not None
        def text(row,key):
            value=row.get(key,""); return "" if value is None else str(value).strip()
        def flag(row,key,default=False):
            raw=text(row,key).lower()
            if raw=="":return default
            if raw in {"1","true","yes","y","on"}:return True
            if raw in {"0","false","no","n","off"}:return False
            raise ValueError(f"{key} must be yes/no or true/false")
        def decimal_value(row,key,required=False,default=None):
            raw=text(row,key)
            if not raw:
                if required: raise ValueError(f"{key} is required")
                return default
            try:return Decimal(raw)
            except InvalidOperation:raise ValueError(f"{key} must be a number")
        def resolve_brand(value):
            if not value:raise ValueError("brand is required")
            qs=Brand.objects.filter(pk=int(value)) if value.isdigit() else Brand.objects.filter(Q(name__iexact=value)|Q(slug__iexact=slugify(value)))
            obj=qs.first()
            if not obj:raise ValueError(f"Brand '{value}' was not found")
            return obj
        def resolve_category(value):
            if not value:raise ValueError("category is required")
            qs=Category.objects.filter(pk=int(value)) if value.isdigit() else Category.objects.filter(Q(name__iexact=value)|Q(slug__iexact=slugify(value)))
            obj=qs.first()
            if not obj:raise ValueError(f"Category '{value}' was not found")
            return obj
        def unique_slug(base,exclude_id=None):
            base=(slugify(base)[:220] or "product"); candidate=base; counter=2
            qs=Product.objects.all()
            if exclude_id:qs=qs.exclude(pk=exclude_id)
            while qs.filter(slug=candidate).exists(): candidate=f"{base[:210]}-{counter}"; counter+=1
            return candidate
        def find_product(row):
            pid=text(row,"product_id")
            if pid.isdigit():
                obj=Product.objects.filter(pk=int(pid)).first()
                if obj:return obj
            sku=text(row,"sku")
            if sku:
                obj=Product.objects.filter(sku__iexact=sku).first()
                if obj:return obj
            slug=text(row,"slug")
            if slug:
                obj=Product.objects.filter(slug__iexact=slug).first()
                if obj:return obj
            return None

        created=updated=variants_created=variants_updated=skipped=0; errors=[]; requested_status={}
        for index,row in enumerate(product_rows,start=2):
            try:
                existing=find_product(row); name=text(row,"name") or (existing.name if existing else "")
                if not name:raise ValueError("name is required")
                ptype=(text(row,"product_type") or (existing.product_type if existing else Product.ProductType.SIMPLE)).lower()
                if ptype not in {Product.ProductType.SIMPLE,Product.ProductType.VARIABLE}:raise ValueError("product_type must be simple or variable")
                if existing and existing.product_type!=ptype:raise ValueError("product_type cannot be changed for an existing product")
                sku=text(row,"sku") or None
                if existing and not has_key(row,"sku"):sku=existing.sku
                if ptype==Product.ProductType.SIMPLE and not sku:raise ValueError("sku is required for simple products")
                duplicate=Product.objects.filter(sku__iexact=sku).exclude(pk=getattr(existing,"pk",None)).first() if sku else None
                if duplicate:raise ValueError(f"SKU '{sku}' already exists")
                status=(text(row,"status") or (existing.status if existing else Product.Status.DRAFT)).lower()
                if status not in {Product.Status.DRAFT,Product.Status.ACTIVE,Product.Status.ARCHIVED}:raise ValueError("status must be draft, active or archived")
                brand=resolve_brand(text(row,"brand") or (str(existing.brand_id) if existing else "")); category=resolve_category(text(row,"category") or (str(existing.category_id) if existing else ""))
                with transaction.atomic():
                    if existing:
                        p=Product.objects.select_for_update().get(pk=existing.pk)
                        p.name=name; p.sku=sku; p.barcode=(text(row,"barcode") or None) if has_key(row,"barcode") else p.barcode; p.brand=brand; p.category=category
                        if has_key(row,"base_price"):p.base_price=decimal_value(row,"base_price",required=True)
                        if has_key(row,"compare_at_price"):p.compare_at_price=decimal_value(row,"compare_at_price")
                        if has_key(row,"cost_price"):p.cost_price=decimal_value(row,"cost_price")
                        if has_key(row,"short_description"):p.short_description=text(row,"short_description")
                        if has_key(row,"description"):p.description=text(row,"description")
                        if has_key(row,"weight"):p.weight=decimal_value(row,"weight")
                        if has_key(row,"tax_class"):p.tax_class=text(row,"tax_class")
                        for key in ("featured","new_arrival","bestseller","trending"):
                            if has_key(row,key):setattr(p,key,flag(row,key,getattr(p,key)))
                        if has_key(row,"slug") and text(row,"slug"):p.slug=unique_slug(text(row,"slug"),p.pk)
                        p.status=status
                        if p.product_type==Product.ProductType.VARIABLE and status==Product.Status.ACTIVE and not p.variants.filter(is_active=True).exists():p.status=Product.Status.DRAFT
                        if p.status==Product.Status.ACTIVE:p.published_at=p.published_at or timezone.now()
                        p.save(); updated+=1
                    else:
                        slug=unique_slug(text(row,"slug") or name)
                        initial_status=Product.Status.DRAFT if ptype==Product.ProductType.VARIABLE and status==Product.Status.ACTIVE else status
                        p=Product.objects.create(name=name,slug=slug,product_type=ptype,sku=sku,barcode=text(row,"barcode") or None,brand=brand,category=category,base_price=decimal_value(row,"base_price",required=True),compare_at_price=decimal_value(row,"compare_at_price"),cost_price=decimal_value(row,"cost_price"),status=initial_status,published_at=timezone.now() if initial_status==Product.Status.ACTIVE else None,short_description=text(row,"short_description"),description=text(row,"description"),weight=decimal_value(row,"weight"),tax_class=text(row,"tax_class"),featured=flag(row,"featured"),new_arrival=flag(row,"new_arrival"),bestseller=flag(row,"bestseller"),trending=flag(row,"trending"))
                        created+=1
                    requested_status[p.pk]=status
            except Exception as exc:
                skipped+=1; errors.append({"sheet":"Products","row":index,"error":str(exc)})

        def find_parent(row):
            pid=text(row,"product_id")
            if pid.isdigit():
                p=Product.objects.filter(pk=int(pid)).first()
                if p:return p
            slug=text(row,"product_slug")
            if slug:
                p=Product.objects.filter(slug__iexact=slug).first()
                if p:return p
            name=text(row,"product_name")
            if name:return Product.objects.filter(name__iexact=name).first()
            return None
        def resolve_attributes(raw):
            raw=(raw or "").strip()
            if not raw:return []
            pieces=[x.strip() for x in raw.replace(";","|").split("|") if x.strip()]; values=[]
            for piece in pieces:
                separator="=" if "=" in piece else ":" if ":" in piece else None
                if not separator:raise ValueError(f"Invalid attribute '{piece}'. Use Attribute=Value")
                attribute_name,value_name=[x.strip() for x in piece.split(separator,1)]
                value=AttributeValue.objects.select_related("attribute").filter(attribute__name__iexact=attribute_name,value__iexact=value_name).first()
                if not value:raise ValueError(f"Attribute value '{attribute_name}={value_name}' was not found")
                values.append(value)
            if len({v.attribute_id for v in values})!=len(values):raise ValueError("Only one value per attribute is allowed")
            return values

        for index,row in enumerate(variant_rows,start=2):
            try:
                parent=find_parent(row)
                if not parent:raise ValueError("Parent product was not found")
                if parent.product_type!=Product.ProductType.VARIABLE:raise ValueError("Variants can only be imported for variable products")
                vid=text(row,"variant_id"); sku=text(row,"sku")
                variant=ProductVariant.objects.filter(pk=int(vid)).first() if vid.isdigit() else None
                if not variant and sku:variant=ProductVariant.objects.filter(sku__iexact=sku).first()
                if variant and variant.product_id!=parent.id:raise ValueError("Existing variant belongs to another product")
                if not sku and not variant:raise ValueError("sku is required")
                duplicate=ProductVariant.objects.filter(sku__iexact=sku).exclude(pk=getattr(variant,"pk",None)).first() if sku else None
                if duplicate:raise ValueError(f"Variant SKU '{sku}' already exists")
                attrs=resolve_attributes(text(row,"attributes")) if has_key(row,"attributes") else None
                with transaction.atomic():
                    if variant:
                        v=ProductVariant.objects.select_for_update().get(pk=variant.pk); v.sku=sku or v.sku
                        if has_key(row,"barcode"):v.barcode=text(row,"barcode") or None
                        if has_key(row,"price_override"):v.price_override=decimal_value(row,"price_override")
                        if has_key(row,"cost_price"):v.cost_price=decimal_value(row,"cost_price")
                        if has_key(row,"weight"):v.weight=decimal_value(row,"weight")
                        if has_key(row,"is_active"):v.is_active=flag(row,"is_active",v.is_active)
                        v.save(); variants_updated+=1
                    else:
                        v=ProductVariant.objects.create(product=parent,sku=sku,barcode=text(row,"barcode") or None,price_override=decimal_value(row,"price_override"),cost_price=decimal_value(row,"cost_price"),weight=decimal_value(row,"weight"),is_active=flag(row,"is_active",True)); variants_created+=1
                    if attrs is not None:v.attributes.set(attrs)
            except Exception as exc:
                skipped+=1; errors.append({"sheet":"Variants","row":index,"error":str(exc)})

        for product_id,status in requested_status.items():
            if status==Product.Status.ACTIVE:
                p=Product.objects.filter(pk=product_id).first()
                if p and (p.product_type==Product.ProductType.SIMPLE or p.variants.filter(is_active=True).exists()):
                    if p.status!=Product.Status.ACTIVE:p.status=Product.Status.ACTIVE;p.published_at=p.published_at or timezone.now();p.save(update_fields=["status","published_at","updated_at"])
        return success({"created":created,"updated":updated,"variants_created":variants_created,"variants_updated":variants_updated,"skipped":skipped,"errors":errors},"Product import complete.")
class AdminVariantViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=VariantAdminSerializer
    def get_permissions(self):
        classes=[CatalogOrderRead] if self.request.method in {"GET","HEAD","OPTIONS"} else [CatalogAdmin]
        return [permission() for permission in classes]
    queryset=(ProductVariant.objects.select_related("product")
        .prefetch_related("attributes__attribute")
        .annotate(admin_available_stock=Coalesce(Sum("stock_item__stocks__available_stock"),Value(0),output_field=IntegerField()))
        .order_by("-id")); filterset_fields=("product","is_active"); search_fields=("sku","barcode","product__name")
class AdminCategoryViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=CategorySerializer; queryset=Category.objects.select_related("parent").all(); search_fields=("name","slug","description"); filterset_fields=("active","parent"); ordering_fields=("name","order","created_at")
class AdminBrandViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=BrandSerializer; queryset=Brand.objects.all(); search_fields=("name","slug","country"); filterset_fields=("active","featured","country"); ordering_fields=("name","created_at")
class AdminImageViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=ProductImageSerializer; parser_classes=[MultiPartParser,FormParser]; queryset=ProductImage.objects.select_related("product","variant").all(); filterset_fields=("product","variant","image_type","is_primary"); search_fields=("alt_text","product__name","variant__sku")

    @transaction.atomic
    def perform_destroy(self,instance):
        product_id,variant_id,was_primary=instance.product_id,instance.variant_id,instance.is_primary
        instance.delete()
        if was_primary:
            replacement=ProductImage.objects.filter(product_id=product_id,variant_id=variant_id).order_by("order","id").first()
            if replacement:
                replacement.is_primary=True
                replacement.save(update_fields=["is_primary","updated_at"])

    @action(detail=True,methods=["post"],url_path="set-primary")
    def primary(self,request,pk=None): return success(ProductImageSerializer(set_primary_image(image=self.get_object()),context={"request":request}).data,"Primary image updated.")
    @action(detail=False,methods=["post"],url_path="reorder")
    def reorder(self,request):
        product=Product.objects.get(pk=request.data.get("product")); reorder_images(product=product,ordered_ids=request.data.get("image_ids",[])); return success(message="Images reordered.")
class BulkImageUploadView(APIView):
    permission_classes=[CatalogAdmin]; parser_classes=[MultiPartParser,FormParser]

    @transaction.atomic
    def post(self,request):
        product=Product.objects.filter(pk=request.data.get("product")).first()
        if not product:
            raise ValidationError({"product":"A valid product is required."})

        variant_id=request.data.get("variant") or None
        variant=ProductVariant.objects.filter(pk=variant_id,product=product).first() if variant_id else None
        if variant_id and not variant:
            raise ValidationError({"variant":"Invalid variant for product."})

        files=request.FILES.getlist("images")
        if not files:
            raise ValidationError({"images":"Select at least one image."})

        primary_index_raw=request.data.get("primary_index")
        primary_index=None
        if primary_index_raw not in (None, ""):
            try:
                primary_index=int(primary_index_raw)
            except (TypeError,ValueError):
                raise ValidationError({"primary_index":"Primary image index must be an integer."})
            if primary_index < 0 or primary_index >= len(files):
                raise ValidationError({"primary_index":"Primary image index is outside the uploaded image list."})

        scope=ProductImage.objects.select_for_update().filter(product=product,variant=variant)
        # If this is the first gallery upload, make the first image the feature image
        # even when the client did not explicitly send primary_index.
        if primary_index is None and not scope.filter(is_primary=True).exists():
            primary_index=0

        if primary_index is not None:
            scope.filter(is_primary=True).update(is_primary=False)

        max_order=scope.aggregate(max_order=Max("order"))["max_order"]
        start_order=(max_order if max_order is not None else -1)+1
        created=[]
        for i,file_obj in enumerate(files):
            created.append(ProductImage.objects.create(
                product=product,
                variant=variant,
                image=file_obj,
                image_type=ProductImage.ImageType.GALLERY,
                alt_text=product.name,
                order=start_order+i,
                is_primary=(i==primary_index),
            ))

        return success(
            ProductImageSerializer(created,many=True,context={"request":request}).data,
            "Images uploaded.",
            201,
        )

class AdminAttributeViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=AttributeAdminSerializer; queryset=Attribute.objects.all().order_by("display_order","name"); search_fields=("name","slug"); ordering_fields=("display_order","name")
class AdminAttributeValueViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=AttributeValueAdminSerializer; queryset=AttributeValue.objects.select_related("attribute").all(); filterset_fields=("attribute",); search_fields=("value","slug")
class AdminClaimViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=ClaimAdminSerializer; queryset=Claim.objects.all(); filterset_fields=("active",); search_fields=("name",)
class AdminProductClaimViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=ProductClaimAdminSerializer; queryset=ProductClaim.objects.select_related("product","claim","reviewed_by").all(); filterset_fields=("product","claim","is_verified","active")
class AdminBeautyProfileViewSet(ModelViewSet): permission_classes=[CatalogAdmin]; serializer_class=BeautyProfileAdminSerializer; queryset=ProductBeautyProfile.objects.prefetch_related("skin_types","hair_types","concerns","ingredients").all(); filterset_fields=("product",)

class AdminSkinTypeViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=SkinTypeAdminSerializer; queryset=SkinType.objects.all().order_by("name"); search_fields=("name","slug")
class AdminHairTypeViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=HairTypeAdminSerializer; queryset=HairType.objects.all().order_by("name"); search_fields=("name","slug")
class AdminConcernViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=ConcernAdminSerializer; queryset=Concern.objects.all().order_by("name"); search_fields=("name","slug","concern_type"); filterset_fields=("concern_type",)
class AdminIngredientViewSet(ModelViewSet):
    permission_classes=[CatalogAdmin]; serializer_class=IngredientAdminSerializer; queryset=Ingredient.objects.all().order_by("name"); search_fields=("name","slug","description")
